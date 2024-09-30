from azure.identity import DefaultAzureCredential
from azure.mgmt.compute import ComputeManagementClient
from azure.mgmt.network import NetworkManagementClient
import traceback
import pandas as pd
import requests
from datetime import datetime,timedelta
import os.path as path
import csv
from openpyxl import Workbook, load_workbook
import subprocess

def get_workbook(new_row,sheet):
    datebefore=datetime.now()-timedelta(days=2)
    today=datebefore.strftime('%d-%m-%y')
    file = "Azure_PROD_"+str(today)+".xlsx"
    workbook=load_workbook(file)
    sheet=workbook[sheet]
    sheet.append(new_row)
    workbook.save(file)

def create_sheet():
    today=datetime.now().strftime('%d-%m-%y')
    file = "Azure_PROD_"+str(today)+".xlsx"
    wb=Workbook()
    sheet=wb.active
    sheet.title='Servers'
    wb['Servers'].append(['Subscription ID','VM ID','IMAGE ID','LAUNCHTIME','STATE','REGION','PRIVATE IP','PUBLIC IP','TYPE','NAME','OwnerEmail','BE','BU','AE','IMAGE NAME','IMAGE CREATION DATE','IMAGEBU','IMAGEBE','IMAGEAE','RELEASE',
            'IMAGE VERSION','POD','CPEVAL','QUALYSEVAL'])
           
    wb.create_sheet('K8s')
    wb['K8s'].append(['Subscripti ID','VM ID','NAME','LOCATION','STATUS','POD','BU','BE','APPENV','OWNEREMAIL','AKS-CLUSTER'])
    file_exists = path.isfile(file)
    wb.save(file)




def get_k8(compute_client,subscription_id):
    print("Inside k8")
    response=compute_client.virtual_machine_scale_sets.list_all()
    print(response)
    #virtual_machine_scale_sets=response.value
    #print(virtual_machine_scale_sets)
    for vmss in response:
        print(vmss.name)
        if "aks" in vmss.name or "AKS" in vmss.name:
            print("vmss: "+str(vmss))
            print("K8 name: "+str(vmss.name))
            resource_grp=vmss.id.split('resourceGroups/')[1].split('/')[0]

            vmss_info = compute_client.virtual_machine_scale_set_vms.list(
             resource_group_name=resource_grp,
             virtual_machine_scale_set_name=vmss.name,
             expand='instanceView'
            )
            print("k8 info: "+str(vmss_info))
            for info in vmss_info:
                print(info)
                vmId=info.vm_id
                status=''
                try:
                    print(info.instance_view.statuses[0])
                    if info.instance_view.statuses[0].code.startswith("PowerState"):
                        status=info.instance_view.statuses[0].display_status
                    exit()
                except:
                    status='Status not present'

                tags=vmss.tags
                tags = {key.upper(): value for key, value in tags.items()}
                pod=''
                bu=''
                be=''
                owneremail=''
                ae=''
                cluster=''
                if "POD" in tags:
                    pod=tags["POD"]
                if "BUSINESSUNIT" in tags:
                    bu=tags["BUSINESSUNIT"]
                if "BUSINESSENTITY" in tags:
                    be=tags["BUSINESSENTITY"]
                if "OWNEREMAIL" in tags:
                    owneremail=tags["OWNEREMAIL"]
                if "APPLICATIONENV" in tags:
                    ae=tags["APPLICATIONENV"]
                if "AKS-CLUSTER" in tags:
                    cluster=tags["AKS-CLUSTER"]
                new_row=[subscription_id,vmId,vmss.name,vmss.location,status,pod,bu,be,ae,owneremail,cluster]
                get_workbook(new_row,'K8s')


def get_vms(compute_client,network_client,subscription_id):

    virtual_machines = compute_client.virtual_machines.list_all()
    #print("virtual_machines: "+str(virtual_machines))
    try:
        for vm in virtual_machines:
            if vm==None:
                continue
            #print(vm)
            print("print virtual machine name : "+str(vm.name))

            resource_group_name = vm.id.split('/')[4]
            virtual_machine_name = vm.name
            #print(vm.id)
            #print("VM name: "+str(virtual_machine_name))
            print("resource group : "+str(resource_group_name))
            # Retrieve the virtual machine

            #resource_group_name='IICS-QA-STG1-CAI'
            #virtual_machine_name='APIM-JENKINS-QA-SLAVE'
            #print(vm.properties.vmId)
            virtual_machine = compute_client.virtual_machines.get(
            resource_group_name,
            virtual_machine_name,
            expand='instanceView'
            )

            #virtual_machine = compute_client.virtual_machines.instance_view(resource_group_name,vm_name=virtual_machine_name)
            print("VM INFO: "+str(virtual_machine))

            # Retrieve the VM image definition ID from the virtual machine
            VM_ID=virtual_machine.vm_id
            print(virtual_machine.vm_id)
            IMAGE_ID=''
            if virtual_machine.storage_profile.image_reference is not None:
                IMAGE_ID=virtual_machine.storage_profile.image_reference.id

            VM_launchtime=virtual_machine.time_created

            VM_launchtime=datetime.strftime(VM_launchtime,"%Y-%m-%d")
            try:
                print(virtual_machine.instance_view.statuses)
                VM_status=virtual_machine.instance_view.statuses[1].display_status
            except:
                VM_status='Status not present'
            VM_Region=virtual_machine.location

            private_ip=''
            public_ips=[]

            #print(" get ip : "+str(virtual_machine.network_profile.network_interfaces))
            #print(" ip : "+str(virtual_machine.network_profile))



            ip = virtual_machine.network_profile.network_interfaces[0].id
            name=ip.split('/')[-1]
            sub=ip.split('/')[4]
            print(' print name : '+str(name))
            print(' print sub : '+str(sub))

            print(" ips : "+str(network_client.network_interfaces.get(sub,name).ip_configurations))

            ips=network_client.network_interfaces.get(sub,name).ip_configurations
            if ips is not None:
                for ip in ips:
                    print(ip.private_ip_address)
                    private_ip=ip.private_ip_address
                    if ip.public_ip_address is not None:
                        public_ip_name=ip.public_ip_address.id.split('/')[-1]
                        try:
                            public_ip=network_client.public_ip_addresses.get(resource_group_name,public_ip_name)
                            print(public_ip.ip_address)
                            if public_ip.ip_address is not None:
                                public_ips.append(public_ip.ip_address)
                        except:
                            public_ips=[]



            VM_type=virtual_machine.hardware_profile.vm_size

            VM_Tags=''
            if virtual_machine.tags is not None:
                VM_Tags=virtual_machine.tags

            print("VM TAGS : "+str(VM_Tags))

            #print("VM TYPE : "+str(VM_type))

            Name=''
            ownermail=''
            vmBU=''
            vmBE=''
            pod=''
            qualyseval=''
            cpeval=''
            appenv=''
            if VM_Tags!='':
                for k,v in VM_Tags.items():
                    if k.upper()=='NAME':
                        Name=v
                    elif k.upper()=="OWNEREMAIL":
                        ownermail=v
                    elif k.upper()=="BUSINESSENTITY":
                        vmBE=v
                    elif k.upper()=="BUSINESSUNIT":
                        vmBU=v
                    elif k.upper()=="POD":
                        pod=v
                    elif k.upper()=="QUALYSEVAL":
                        qualyseval=v
                    elif k.upper()=='CPEVAL':
                        cpeval=v
                    elif k.upper()=="APPLICATIONENV":
                        appenv=v
            #print(Name)
            #print(ownermail)
            #print(vmBE)
            #print(vmBU)

            vm_image_definition_id = virtual_machine.storage_profile.image_reference
            print(vm_image_definition_id)

            #print("Image: "+str(virtual_machine.storage_profile))
            image_definition_name=''
            image_version=''
            image_resource_grp=''
            image_gallery=''
            image_creation_date=''
            amiBU=''
            amiBE=''
            amiAE=''
            release=''

            print(vm_image_definition_id)
            if vm_image_definition_id is not None and  vm_image_definition_id.id is not None:
                if 'images/' in vm_image_definition_id.id:
                    image_definition_name = vm_image_definition_id.id.split('images/')[1].split('/')[0]
                if 'versions/' in vm_image_definition_id.id:
                    image_version = vm_image_definition_id.id.split('versions/')[1].split('/')[0]
                if 'resourceGroups/' in vm_image_definition_id.id:
                    image_resource_grp= vm_image_definition_id.id.split('resourceGroups/')[1].split('/')[0]
                if 'galleries/' in vm_image_definition_id.id:
                    image_gallery=vm_image_definition_id.id.split('galleries/')[1].split('/')[0]
                if 'subscriptions/' in vm_image_definition_id.id:
                    image_subscription=vm_image_definition_id.id.split('subscriptions/')[1].split('/')[0]
            # Print the image definition name and image creation date
                print(f"Image Definition Name: {image_definition_name}")
                print(f"Image Version: {image_version}")
                print(f"Image Resource grp: {image_resource_grp}")
                print(f"Image Gallery: {image_gallery}")

                images=None

                if image_definition_name!='' and image_version!='' and image_resource_grp!='' and image_gallery!='' and image_subscription!='':
                    try:

                        compute_client2 = ComputeManagementClient(credential=credential, subscription_id=image_subscription,api_version='2022-03-03')

                        images = compute_client2.gallery_image_versions.get(

                            resource_group_name=image_resource_grp,
                            gallery_name=image_gallery,
                            gallery_image_name=image_definition_name,
                            gallery_image_version_name=image_version

                        )
                        print(images)
                    except:
                        print('Image not there1')
                        print(traceback.print_exc())



                    # Create a list to hold image details
                    image_list = []



                    if images is not None:
                        image_info = {

                            "location": images.location,

                            "publisher":images.publishing_profile.published_date,

                            "tags": images.tags,

                        }

                        print(image_info['tags'])

                        print(" GETTING AMI DETAILS ")


                        image_creation_date=image_info['publisher']



                        image_creation_date = image_creation_date.strftime("%Y-%m-%d")
                        print(image_creation_date)
                        for k,v in image_info['tags'].items():

                            if k.upper()=="BUSINESSENTITY":
                                amiBE=v
                            elif k.upper()=="BUSINESSUNIT":
                                amiBU=v
                            elif k.upper()=="APPLICATIONENV":
                                amiAE=v
                            elif "RELEASE" in k.upper():
                                release=v

                    print(" AMI DETAILS FINISHED ")

            new_row=[subscription_id,VM_ID,IMAGE_ID,VM_launchtime,VM_status,VM_Region,private_ip,','.join(public_ips),VM_type,virtual_machine_name,ownermail,vmBE,vmBU,appenv,image_definition_name,
                        image_creation_date,amiBU,amiBE,amiAE,release,image_version,pod,cpeval,qualyseval]
            #new_row=[subscription_id,VM_ID,IMAGE_ID,VM_launchtime,VM_status,VM_Region,Name,ownermail,vmBE,vmBU,appenv,image_definition_name,
                        #image_creation_date,amiBU,amiBE,amiAE,release,image_version,pod,cpeval,qualyseval]
            print(new_row)
            get_workbook(new_row,'Servers')

    except:
        print('Image not there2')
        print(traceback.print_exc())






create_sheet()



#add list of subscriptionId
subscription_list = []


for subscription_id in subscription_list:

    # Set your Azure credentials
    subprocess.run("az account set --subscription "+subscription_id,shell=True)
    credential = DefaultAzureCredential()


    compute_client = ComputeManagementClient(credential=credential, subscription_id=subscription_id)
    network_client=NetworkManagementClient(credential,subscription_id)
    get_k8(compute_client,subscription_id)
    get_vms(compute_client,network_client,subscription_id)
    #get_vms(compute_client,subscription_id)
